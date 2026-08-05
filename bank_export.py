#!/usr/bin/env python3
"""
bank_export.py — Exportador de Extratos Bancários
Formatos: CSV, TXT, JSON, XML, HTML, Markdown, Excel (xlsx), PDF
Uso CLI: python3 bank_export.py <arquivo.txt> [diretorio_saida]
"""

from __future__ import annotations

import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from fpdf import FPDF

    HAS_PDF = True
except ImportError:
    HAS_PDF = False


HEADERS = ["Data", "Hora", "Tipo", "Valor (R$)", "Status", "Descricao"]

_TIPO_LABEL: dict[str, str] = {
    "DEP": "Deposito",
    "SAQ": "Saque",
    "TED": "TED",
    "DOC": "DOC",
    "PIX": "PIX",
    "PAG": "Pagamento",
    "EMP": "Emprestimo",
    "PMT": "Parcela",
    "FAT": "Fatura",
    "AGN": "Agendado",
    "TAR": "Tarifa",
    "REN": "Rendimento",
    "BOL": "Boleto",
    "TRF": "Transferencia",
}

_STATUS_LABEL: dict[str, str] = {
    "E": "Efetivada",
    "P": "Pendente",
    "X": "Estornada",
    "C": "Cancelada",
    "F": "Falhou",
}

_PAT_PIPE = re.compile(
    r"(\d{8})"
    r"(?:\s*[|\s]\s*(\d{6}))?"
    r"\s*[|\s]\s*([A-Z]{2,3})"
    r".*?R\$\s*([\d.,]+-?)"
    r".*?([PEXCF])\s*$",
    re.IGNORECASE,
)

_PAT_FIELDS = re.compile(
    r"(\d{8})"
    r"[^\w]*([A-Z]{2,3})[^\w]*"
    r"R\$\s*([\d.,]+-?)"
    r".*?([PEXCF])\s*$",
    re.IGNORECASE,
)



def parse_extrato_lines(text: str) -> list[dict[str, Any]]:
    """Converte linhas brutas do COBOL em lista de dicts de transação."""
    records: list[dict[str, Any]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            tipo = ""
            for p in parts:
                if p.upper() in _TIPO_LABEL:
                    tipo = p.upper()
                    break
            if tipo:
                valor = next((p for p in parts if re.search(r"\d+[.,]\d{2}", p)), "")
                data = next((p for p in parts if re.fullmatch(r"\d{8}", p)), "")
                status = next((p for p in parts if p.upper() in _STATUS_LABEL), "")
                descr = next((p for p in parts if len(p) > 4 and p not in (tipo, data, status, valor)), line)
                records.append(
                    {
                        "data": data,
                        "hora": "",
                        "tipo": tipo,
                        "valor": valor,
                        "status": status.upper(),
                        "descricao": descr,
                    }
                )
                continue

        m = _PAT_PIPE.search(line) or _PAT_FIELDS.search(line)
        if m:
            g = m.groups()
            records.append(
                {
                    "data": g[0],
                    "hora": g[1] if len(g) > 4 else "",
                    "tipo": (g[2] if len(g) > 4 else g[1]).upper(),
                    "valor": g[3] if len(g) > 4 else g[2],
                    "status": (g[4] if len(g) > 4 else g[3]).upper(),
                    "descricao": line,
                }
            )

    return records


def _humanize(records: list[dict[str, Any]]) -> list[list[str]]:
    return [
        [
            _fmt_date(r["data"]),
            _fmt_time(r["hora"]),
            _TIPO_LABEL.get(r["tipo"], r["tipo"]),
            r["valor"],
            _STATUS_LABEL.get(r["status"], r["status"]),
            r["descricao"],
        ]
        for r in records
    ]


def _fmt_date(d: str) -> str:
    try:
        return f"{d[6:8]}/{d[4:6]}/{d[:4]}" if len(d) == 8 else d
    except Exception:
        return d


def _fmt_time(t: str) -> str:
    try:
        return f"{t[:2]}:{t[2:4]}:{t[4:6]}" if len(t) == 6 else t
    except Exception:
        return t




def export_csv(records: list[dict], path: Path) -> bool:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(HEADERS)
        w.writerows(_humanize(records))
    return True


def export_txt(records: list[dict], path: Path, raw: str = "") -> bool:
    with open(path, "w", encoding="utf-8") as f:
        f.write("=" * 100 + "\n")
        f.write("  EXTRATO BANCARIO — Sistema Bancario COBOL\n")
        f.write(
            f"  Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            f"   |   {len(records)} transacao(oes)\n"
        )
        f.write("=" * 100 + "\n\n")
        if raw:
            f.write(raw)
        else:
            col_w = [12, 10, 16, 18, 12, 40]
            header = "".join(h.ljust(w) for h, w in zip(HEADERS, col_w))
            f.write(header + "\n")
            f.write("-" * sum(col_w) + "\n")
            for row in _humanize(records):
                f.write("".join(str(v).ljust(w) for v, w in zip(row, col_w)) + "\n")
    return True


def export_json(records: list[dict], path: Path) -> bool:
    payload = {
        "exportado_em": datetime.now().isoformat(),
        "sistema": "Banco COBOL v2.0",
        "total_transacoes": len(records),
        "transacoes": records,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return True


def export_xml(records: list[dict], path: Path) -> bool:
    ts = datetime.now().isoformat()
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        f'<extrato exportado_em="{ts}" total="{len(records)}" sistema="Banco COBOL v2.0">',
    ]
    for r in records:
        lines.append("  <transacao>")
        for k, v in r.items():
            safe = str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            lines.append(f"    <{k}>{safe}</{k}>")
        lines.append("  </transacao>")
    lines.append("</extrato>")
    path.write_text("\n".join(lines), encoding="utf-8")
    return True


def export_html(records: list[dict], path: Path) -> bool:
    rows_html = "\n".join(
        "    <tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>" for row in _humanize(records)
    )
    header_html = "".join(f"<th>{h}</th>" for h in HEADERS)
    ts = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Extrato Bancário</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Arial,sans-serif;background:#f0f4f8;padding:24px;color:#222}}
  .card{{background:#fff;border-radius:8px;box-shadow:0 2px 12px rgba(0,0,0,.1);padding:24px;max-width:1200px;margin:0 auto}}
  h1{{color:#1a3c6e;margin-bottom:4px;font-size:1.6rem}}
  .meta{{color:#666;font-size:.9rem;margin-bottom:20px}}
  table{{width:100%;border-collapse:collapse;font-size:.9rem}}
  thead tr{{background:#1a3c6e;color:#fff}}
  th{{padding:10px 14px;text-align:left;white-space:nowrap}}
  td{{padding:8px 14px;border-bottom:1px solid #e8edf2}}
  tbody tr:hover{{background:#f0f4ff}}
  tbody tr:nth-child(even){{background:#f8fafc}}
  tbody tr:nth-child(even):hover{{background:#e8f0ff}}
  .badge{{display:inline-block;padding:2px 8px;border-radius:12px;font-size:.8rem;font-weight:600}}
  .footer{{margin-top:16px;text-align:right;color:#999;font-size:.8rem}}
</style>
</head>
<body>
<div class="card">
  <h1>&#128196; Extrato Bancário</h1>
  <p class="meta">Exportado em: <strong>{ts}</strong> &nbsp;|&nbsp; <strong>{len(records)}</strong> transação(ões)</p>
  <table>
    <thead><tr>{header_html}</tr></thead>
    <tbody>
{rows_html}
    </tbody>
  </table>
  <div class="footer">Sistema Bancário COBOL v2.0 — Exportação automática</div>
</div>
</body>
</html>"""
    path.write_text(html, encoding="utf-8")
    return True


def export_markdown(records: list[dict], path: Path) -> bool:
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    lines = [
        "# Extrato Bancário",
        "",
        f"> **Sistema:** Banco COBOL v2.0  |  **Exportado:** {ts}  |  **Transações:** {len(records)}",
        "",
        "| " + " | ".join(HEADERS) + " |",
        "|" + "|".join(":---" for _ in HEADERS) + "|",
    ]
    for row in _humanize(records):
        lines.append("| " + " | ".join(str(v).replace("|", "\\|") for v in row) + " |")
    path.write_text("\n".join(lines), encoding="utf-8")
    return True


def export_excel(records: list[dict], path: Path) -> bool:
    if not HAS_EXCEL:
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato"

    header_fill = PatternFill("solid", fgColor="1A3C6E")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", fgColor="F0F4FF")
    for idx, row in enumerate(_humanize(records), start=2):
        ws.append(row)
        if idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = alt_fill

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 60)

    ws.freeze_panes = "A2"

    ms = wb.create_sheet("Info")
    ms["A1"] = "Sistema"
    ms["B1"] = "Banco COBOL v2.0"
    ms["A2"] = "Exportado em"
    ms["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ms["A3"] = "Total de transacoes"
    ms["B3"] = len(records)

    wb.save(path)
    return True


def export_pdf(records: list[dict], path: Path) -> bool:
    if not HAS_PDF:
        return False

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=12)

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(26, 60, 110)
    pdf.cell(0, 10, "Extrato Bancário", ln=True, align="C")

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 100, 100)
    ts = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    pdf.cell(
        0,
        6,
        f"Sistema Bancário COBOL v2.0   |   Exportado: {ts}   |   {len(records)} transações",
        ln=True,
        align="C",
    )
    pdf.ln(4)

    col_w = [26, 18, 24, 30, 24, 0]
    page_w = 277 - 20
    col_w[-1] = page_w - sum(col_w[:-1])

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(26, 60, 110)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(col_w, HEADERS):
        pdf.cell(w, 8, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for idx, row in enumerate(_humanize(records)):
        if idx % 2 == 0:
            pdf.set_fill_color(240, 244, 255)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 30, 30)
        for w, cell in zip(col_w, row):
            pdf.cell(w, 7, str(cell)[:35], border=1, fill=True)
        pdf.ln()

    pdf.set_y(-15)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f"Banco COBOL v2.0 — Página {pdf.page_no()}", align="C")

    pdf.output(str(path))
    return True



EXPORTERS: dict[str, Any] = {
    "csv": export_csv,
    "txt": export_txt,
    "json": export_json,
    "xml": export_xml,
    "html": export_html,
    "md": export_markdown,
    "xlsx": export_excel,
    "pdf": export_pdf,
}

AVAILABLE_FORMATS = list(EXPORTERS.keys())


def export_formats(
    records: list[dict],
    formats: list[str],
    out_dir: Path,
    stem: str = "extrato",
    raw_text: str = "",
) -> dict[str, bool]:
    """Exporta para os formatos especificados. Retorna {fmt: sucesso}."""
    out_dir.mkdir(parents=True, exist_ok=True)
    results: dict[str, bool] = {}
    for fmt in formats:
        fn = EXPORTERS.get(fmt)
        if fn is None:
            results[fmt] = False
            continue
        out_path = out_dir / f"{stem}.{fmt}"
        try:
            if fmt == "txt":
                ok = fn(records, out_path, raw=raw_text)
            else:
                ok = fn(records, out_path)
            results[fmt] = ok is not False
        except Exception as exc:
            results[fmt] = False
            print(f"[EXPORT] Erro ao gerar {fmt}: {exc}", file=sys.stderr)
    return results


def export_all(
    records: list[dict],
    out_dir: Path,
    stem: str = "extrato",
    raw_text: str = "",
) -> dict[str, bool]:
    return export_formats(records, AVAILABLE_FORMATS, out_dir, stem, raw_text)




def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: python3 bank_export.py <extrato.txt> [diretorio_saida] [formato1,formato2,...]")
        print(f"Formatos disponíveis: {', '.join(AVAILABLE_FORMATS)}")
        sys.exit(1)

    raw = Path(sys.argv[1]).read_text(encoding="utf-8", errors="replace")
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("exports")
    fmts_arg = sys.argv[3].split(",") if len(sys.argv) > 3 else AVAILABLE_FORMATS
    fmts = [f.strip().lower() for f in fmts_arg if f.strip().lower() in EXPORTERS]

    records = parse_extrato_lines(raw)
    print(f"Transações encontradas: {len(records)}")

    stem = f"extrato_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    results = export_formats(records, fmts, out_dir, stem, raw_text=raw)

    for fmt, ok in results.items():
        icon = "✓" if ok else "✗"
        print(f"  {icon} {fmt.upper():6}  →  {out_dir / f'{stem}.{fmt}'}")


if __name__ == "__main__":
    main()
